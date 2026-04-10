"""
Генератор коммерческих предложений.

Принимает input.json или input.xlsx, рассчитывает объёмы и стоимость,
генерирует data.json и компилирует PDF через Typst.

Использование:
    python generate.py input.json
    python generate.py input.xlsx
    python generate.py input.json -o КП_Нива.pdf
"""

import argparse
import json
import subprocess
import sys
from pathlib import Path

ROOT = Path(__file__).parent
DATA_START_ROW = 13  # первая строка данных в XLSX таблице сортов

PAIN_MAP = {"засуха": "засуха", "болезни": "болезни", "качество": "качество"}


# ══════════════════════════════════════════════════════════════
# Чтение входных данных
# ══════════════════════════════════════════════════════════════

def read_json_input(path: str) -> dict:
    """
    Читает input.json — минимальный формат для AI/скрипта.

    Схема input.json:
    {
      "client_name": "ООО Нива",
      "region": "Оренбургская область",
      "client_email": "",          // опционально
      "client_phone": "",          // опционально
      "manager_name": "",          // опционально
      "manager_email": "",         // опционально
      "manager_phone": "",         // опционально
      "entries": [
        {"variety": "ДФ 2020", "area_ha": 1000, "pain": "Засуха"},
        {"variety": "Володя",  "area_ha": 500,  "pain": "Качество"}
      ]
    }
    """
    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    # Валидация
    for field in ("client_name", "region", "entries"):
        if field not in raw or not raw[field]:
            print(f"ОШИБКА: отсутствует обязательное поле '{field}' в input.json")
            sys.exit(1)

    entries = []
    for i, e in enumerate(raw["entries"], 1):
        variety = str(e.get("variety", "")).strip()
        if not variety:
            print(f"ОШИБКА: entries[{i}]: нет сорта")
            sys.exit(1)

        try:
            area = float(str(e.get("area_ha", "")).replace(" ", "").replace(",", "."))
        except (ValueError, TypeError):
            print(f"ОШИБКА: entries[{i}]: площадь не число ({e.get('area_ha')})")
            sys.exit(1)

        pain_str = str(e.get("pain", "")).strip().lower()
        if pain_str not in PAIN_MAP:
            print(f"ОШИБКА: entries[{i}]: неизвестная боль '{e.get('pain')}'. Допустимо: Засуха, Болезни, Качество")
            sys.exit(1)

        entries.append({
            "variety": variety,
            "area_ha": area,
            "pain": PAIN_MAP[pain_str],
        })

    return {
        "header": {
            "client_name": str(raw["client_name"]).strip(),
            "region": str(raw["region"]).strip(),
            "client_email": str(raw.get("client_email", "") or ""),
            "client_phone": str(raw.get("client_phone", "") or ""),
            "manager_name": str(raw.get("manager_name", "") or ""),
            "manager_email": str(raw.get("manager_email", "") or ""),
            "manager_phone": str(raw.get("manager_phone", "") or ""),
        },
        "entries": entries,
    }


def read_xlsx_input(path: str) -> dict:
    """Читает XLSX: шапку (клиент + менеджер) и таблицу сортов."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    def cell(r, c):
        v = ws.cell(row=r, column=c).value
        return str(v).strip() if v is not None else ""

    header = {
        "client_name": cell(3, 2),
        "region": cell(4, 2),
        "client_email": cell(5, 2),
        "client_phone": cell(6, 2),
        "manager_name": cell(7, 2),
        "manager_email": cell(8, 2),
        "manager_phone": cell(9, 2),
    }

    if not header["client_name"]:
        print("ОШИБКА: не заполнено имя клиента (B3)")
        sys.exit(1)
    if not header["region"]:
        print("ОШИБКА: не заполнен регион (B4)")
        sys.exit(1)

    entries = []
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        variety = ws.cell(row=row_idx, column=2).value
        if not variety or not str(variety).strip():
            continue

        area_raw = ws.cell(row=row_idx, column=3).value
        pain_raw = ws.cell(row=row_idx, column=4).value

        try:
            area = float(str(area_raw).replace(" ", "").replace(",", "."))
        except (ValueError, TypeError):
            print(f"ОШИБКА: строка {row_idx}: площадь не число ({area_raw})")
            sys.exit(1)

        pain_str = str(pain_raw or "").strip().lower()
        if pain_str not in PAIN_MAP:
            print(f"ОШИБКА: строка {row_idx}: неизвестная боль ({pain_raw})")
            sys.exit(1)

        entries.append({
            "variety": str(variety).strip(),
            "area_ha": area,
            "pain": PAIN_MAP[pain_str],
        })

    if not entries:
        print("ОШИБКА: таблица сортов пуста")
        sys.exit(1)

    return {"header": header, "entries": entries}


def read_input(path: str) -> dict:
    """Автоопределение формата по расширению: .json или .xlsx"""
    p = Path(path)
    if p.suffix.lower() == ".json":
        return read_json_input(path)
    elif p.suffix.lower() == ".xlsx":
        return read_xlsx_input(path)
    else:
        print(f"ОШИБКА: неподдерживаемый формат '{p.suffix}'. Используйте .json или .xlsx")
        sys.exit(1)


# ══════════════════════════════════════════════════════════════
# Справочники
# ══════════════════════════════════════════════════════════════

def load_varieties() -> dict:
    with open(ROOT / "ref" / "varieties.json", "r", encoding="utf-8") as f:
        return json.load(f)


def load_technology() -> dict:
    with open(ROOT / "ref" / "technology.json", "r", encoding="utf-8") as f:
        return json.load(f)


# ══════════════════════════════════════════════════════════════
# Расчёт
# ══════════════════════════════════════════════════════════════

def calculate_volumes(tech: dict, area_ha: float) -> dict:
    seeding_rate = tech["seeding_rate_kg_per_ha"]
    seed_tonnes = (area_ha * seeding_rate) / 1000

    seasons = []
    grand_total = 0

    for season in tech["seasons"]:
        treatments = []
        for treatment in season["treatments"]:
            products = []
            for p in treatment["products"]:
                rate = p["rate"]
                if p.get("rate_basis") == "seed":
                    volume = round(rate * seed_tonnes, 2)
                else:
                    volume = round(rate * area_ha, 2)

                price = p["price_per_unit"]
                cost = round(volume * price, 2)
                grand_total += cost

                products.append({
                    "name": p["name"],
                    "category": p["category"],
                    "rate": rate,
                    "unit": p["unit"],
                    "volume": volume,
                    "price": price,
                    "cost": cost,
                })
            treatments.append({
                "num": treatment["num"],
                "phase": treatment["phase"],
                "products": products,
            })
        seasons.append({
            "name": season["name"],
            "treatments": treatments,
        })

    seed_price = 25000
    seed_cost = round(seed_tonnes * seed_price, 2)
    grand_total += seed_cost

    return {
        "seasons": seasons,
        "seeds": {
            "variety": "",
            "volume_tonnes": round(seed_tonnes, 2),
            "price_per_tonne": seed_price,
            "cost": seed_cost,
        },
        "grand_total": round(grand_total, 2),
    }


def build_data(input_data: dict, seeds_only: bool = False) -> dict:
    """Собирает полный data.json с массивом entries."""
    header = input_data["header"]
    all_varieties = load_varieties()
    tech = load_technology()

    available = ", ".join(all_varieties.keys())
    entries = []
    total_area = 0
    total_cost = 0

    for e in input_data["entries"]:
        vname = e["variety"]
        if vname not in all_varieties:
            print(f"ОШИБКА: сорт '{vname}' не найден. Доступные: {available}")
            sys.exit(1)

        v = all_varieties[vname]
        area = e["area_ha"]
        pain = e["pain"]
        if seeds_only:
            # Только семена — без агрохимии
            seeding_rate = tech["seeding_rate_kg_per_ha"]
            seed_tonnes = round((area * seeding_rate) / 1000, 2)
            seed_price = 25000
            seed_cost = round(seed_tonnes * seed_price, 2)
            calc = {
                "seasons": [],
                "seeds": {
                    "variety": vname,
                    "volume_tonnes": seed_tonnes,
                    "price_per_tonne": seed_price,
                    "cost": seed_cost,
                },
                "grand_total": seed_cost,
            }
        else:
            calc = calculate_volumes(tech, area)
            calc["seeds"]["variety"] = vname

        pain_argument = v.get("pain_arguments", {}).get(pain, "")
        resistance = [
            {"disease": d.replace("_", " "), "level": l}
            for d, l in v.get("disease_resistance", {}).items()
        ]

        entries.append({
            "variety": {
                "name": vname,
                "biotype": v.get("biotype", ""),
                "quality_class": v.get("quality_class", ""),
                "yield_avg": v.get("yield_avg"),
                "yield_max": v.get("yield_max"),
                "yield_production": v.get("yield_production"),
                "protein_percent": v.get("protein_percent"),
                "gluten_percent": v.get("gluten_percent"),
                "key_advantages": v.get("key_advantages", []),
                "registry_since": v.get("registry_since"),
                "region_names": v.get("region_names", []),
            },
            "area_ha": area,
            "pain": pain,
            "pain_label": {"засуха": "Засуха", "болезни": "Болезни", "качество": "Качество"}[pain],
            "pain_argument": pain_argument,
            "resistance": resistance,
            "calculation": calc,
        })

        total_area += area
        total_cost += calc["grand_total"]

    return {
        "client_name": header["client_name"],
        "region": header["region"],
        "manager_name": header["manager_name"],
        "email": header["client_email"] or header["manager_email"],
        "phone": header["client_phone"] or header["manager_phone"],
        "technology": {
            "title": tech["title"],
            "crop": tech["crop"],
            "seeding_rate": tech["seeding_rate_kg_per_ha"],
        },
        "entries": entries,
        "total_area": total_area,
        "total_cost": round(total_cost, 2),
    }


# ══════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Генератор КП (input.json или input.xlsx)")
    parser.add_argument("input", help="Путь к input.json или input.xlsx")
    parser.add_argument("-o", "--output", default=None, help="Имя выходного PDF")
    parser.add_argument("--seeds-only", action="store_true", help="Только семена (без агрохимии), шаблон seeds.typ")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ОШИБКА: файл не найден: {input_path}")
        sys.exit(1)

    print(f"Читаю {input_path}...")
    input_data = read_input(str(input_path))

    n = len(input_data["entries"])
    print(f"Клиент: {input_data['header']['client_name']}")
    print(f"Сортов: {n}")
    for e in input_data["entries"]:
        print(f"  - {e['variety']} | {e['area_ha']} га | {e['pain']}")

    print("Рассчитываю" + (" (только семена)..." if args.seeds_only else "..."))
    data = build_data(input_data, seeds_only=args.seeds_only)

    data_path = ROOT / "data.json"
    with open(data_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    output_name = args.output
    if not output_name:
        client = data["client_name"].replace('"', "").replace(" ", "_").replace("/", "_")[:30]
        output_name = f"КП_{client}.pdf"

    output_path = ROOT / output_name
    typst_main = ROOT / ("seeds.typ" if args.seeds_only else "main.typ")

    print(f"Компилирую -> {output_path}...")
    result = subprocess.run(
        ["typst", "compile", str(typst_main), str(output_path)],
        capture_output=True,
        text=True,
    )

    if result.returncode != 0:
        print(f"ОШИБКА Typst:\n{result.stderr}")
        sys.exit(1)

    cost_per_ha = round(data["total_cost"] / data["total_area"])
    print(f"Готово: {output_path}")
    print(f"Итого: {data['total_area']:.0f} га, {data['total_cost']:,.0f} руб. ({cost_per_ha:,} руб/га)")


if __name__ == "__main__":
    main()
