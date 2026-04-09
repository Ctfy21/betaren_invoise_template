"""
Создаёт input_template.xlsx — шаблон для заполнения продажником.
Шапка: данные клиента + менеджер.
Таблица: строки с сортами — каждая строка = отдельный сорт/поле клиента.
Один файл = один клиент = один PDF.
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).parent
MAX_ROWS = 30  # максимум сортов в одном КП


def create_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "КП"

    # ── Стили ──
    title_font = Font(name="Arial", size=14, bold=True, color="125488")
    label_font = Font(name="Arial", size=10, bold=True, color="1D646A")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    hint_font = Font(name="Arial", size=9, color="888888", italic=True)
    teal_fill = PatternFill(start_color="1D646A", end_color="1D646A", fill_type="solid")
    input_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
    alt_fill = PatternFill(start_color="F4F8F9", end_color="F4F8F9", fill_type="solid")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ── Ширины столбцов ──
    widths = {"A": 20, "B": 32, "C": 14, "D": 16, "E": 26}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Заголовок ──
    ws.merge_cells("A1:E1")
    ws["A1"] = "Коммерческое предложение — Щёлково Агрохим"
    ws["A1"].font = title_font

    # ── Шапка: клиент + менеджер ──
    header_fields = [
        (3, "Имя клиента", 'ООО «Название»'),
        (4, "Регион", "Оренбургская область"),
        (5, "Email клиента", "client@example.com"),
        (6, "Телефон клиента", "+7 (XXX) XXX-XX-XX"),
        (7, "Имя менеджера", "Иванов Иван Иванович"),
        (8, "Email менеджера", "manager@betaren.ru"),
        (9, "Телефон менеджера", "+7 (495) 777-84-89"),
    ]
    for row, label, hint in header_fields:
        cl = ws.cell(row=row, column=1, value=label)
        cl.font = label_font
        cl.alignment = Alignment(vertical="center")

        cv = ws.cell(row=row, column=2)
        cv.fill = input_fill
        cv.border = thin
        cv.alignment = Alignment(vertical="center")
        cv.number_format = "@"

        ch = ws.cell(row=row, column=3, value=hint)
        ch.font = hint_font

    # ── Разделитель ──
    ws.cell(row=11, column=1, value="Сорта и площади:").font = Font(
        name="Arial", size=11, bold=True, color="125488"
    )

    # ── Заголовок таблицы (строка 12) ──
    TABLE_START = 12
    table_cols = [
        ("A", "№"),
        ("B", "Сорт"),
        ("C", "Площадь (га)"),
        ("D", "Главная боль"),
        ("E", "Комментарий"),
    ]
    for col_letter, title in table_cols:
        cell = ws[f"{col_letter}{TABLE_START}"]
        cell.value = title
        cell.font = header_font
        cell.fill = teal_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin

    # ── Строки данных ──
    DATA_START = TABLE_START + 1
    for i in range(MAX_ROWS):
        row = DATA_START + i
        bg = alt_fill if i % 2 == 1 else input_fill

        # №
        cn = ws.cell(row=row, column=1, value=i + 1)
        cn.font = Font(name="Arial", size=9, color="888888")
        cn.alignment = Alignment(horizontal="center", vertical="center")
        cn.border = thin

        # Сорт (B), Площадь (C), Боль (D), Комментарий (E)
        for col_idx in range(2, 6):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = bg
            cell.border = thin
            cell.alignment = Alignment(vertical="center")
            if col_idx == 3:
                cell.number_format = "#,##0"
            else:
                cell.number_format = "@"

    # ── Валидация: сорт ──
    with open(ROOT / "ref" / "varieties.json", "r", encoding="utf-8") as f:
        varieties = json.load(f)
    variety_formula = '"' + ",".join(varieties.keys()) + '"'

    dv_variety = DataValidation(
        type="list",
        formula1=variety_formula,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Неверный сорт",
        error="Выберите сорт из списка",
        showInputMessage=True,
        promptTitle="Сорт",
        prompt="Выберите сорт озимой пшеницы",
    )
    ws.add_data_validation(dv_variety)
    dv_variety.add(f"B{DATA_START}:B{DATA_START + MAX_ROWS - 1}")

    # ── Валидация: боль ──
    dv_pain = DataValidation(
        type="list",
        formula1='"Засуха,Болезни,Качество"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Неверное значение",
        error="Выберите: Засуха, Болезни или Качество",
        showInputMessage=True,
        promptTitle="Боль",
        prompt="Засуха / Болезни / Качество",
    )
    ws.add_data_validation(dv_pain)
    dv_pain.add(f"D{DATA_START}:D{DATA_START + MAX_ROWS - 1}")

    # ── Подсказки ──
    hr = DATA_START + MAX_ROWS + 1
    ws.cell(row=hr, column=1, value="Заполняйте строки сверху вниз. Пустые строки пропускаются.").font = hint_font
    ws.cell(row=hr + 1, column=1, value="Один файл = один клиент. Каждая строка = отдельный сорт/поле.").font = hint_font
    ws.cell(row=hr + 2, column=1, value="Генерация: python generate.py input.xlsx").font = hint_font

    # ── Закрепить шапку таблицы ──
    ws.freeze_panes = f"A{DATA_START}"

    # ── Сохранение ──
    out_path = ROOT / "input_template.xlsx"
    wb.save(out_path)
    print(f"Шаблон создан: {out_path}")


if __name__ == "__main__":
    create_template()
